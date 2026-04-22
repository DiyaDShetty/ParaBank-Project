import sys
import os
import time
import pytest
from selenium import webdriver
from openpyxl import load_workbook

# ✅ Fix import path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from pages.invalid_login_login_page import LoginPage


# 🔹 Setup WebDriver
def setup_driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://parabank.parasoft.com/parabank/index.htm")
    return driver


# 🔹 Read Excel Data
def get_test_data():
    file_path = os.path.join(
        os.path.dirname(__file__),
        "..",
        "testdata",
        "Parabank_Invalid_Login_TestCases_Desc.xlsx"
    )

    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        description = row[0]
        test_case = row[1]
        username = row[2] or ""   # handle empty cells
        password = row[3] or ""
        expected = row[4]

        data.append((description, test_case, username, password, expected))

    return data


# 🔹 Generate readable test IDs
test_data = get_test_data()
test_ids = [row[1] for row in test_data]  # using "Test Cases" column


# 🔹 Data-driven Test
@pytest.mark.parametrize(
    "description,test_case,username,password,expected",
    test_data,
    ids=test_ids
)
def test_invalid_login(description, test_case, username, password, expected):
    driver = setup_driver()

    try:
        login = LoginPage(driver)
        login.login(username, password)

        time.sleep(2)

        error_title = login.get_error_title()

        print("\n----------------------------")
        print("Description:", description)
        print("Test Case:", test_case)
        print("Username:", username)
        print("Password:", password)
        print("Expected:", expected)
        print("Actual:", error_title)

        # ✅ Assertion

        print("Error Title:", error_title)

        assert error_title == "Error!"

    finally:
        time.sleep(1)
        driver.quit()